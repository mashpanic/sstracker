#!/usr/bin/env python3
"""
mt2_extract_roster.py — dump every enemy/character to roster.json (+ roster.xlsx).

Reads CharacterData objects from sharedassets0.assets and resolves display names
from resources.assets. Filters cutscene portraits (DialogueCharacter_*).

Columns: name, internal, set, hp, atk, subtypes, kind, is_boss, is_support, is_snack.

Usage:  python3 mt2_extract_roster.py [DATA_DIR] [--out OUTDIR] [--no-xlsx]
"""
import sys, os, re, json
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mt2_lib import parse_assets, iter_characters, Localizer, find_data_dir, EXCLUDED_INTERNALS

def classify(internal, subs, hp, atk):
    s=(' '.join(subs)+' '+internal).lower()
    if internal.startswith('DialogueCharacter'): return 'Cutscene'
    if 'snack' in s: return 'Snack/Pickup'
    if any(k in s for k in ['treasure','gold','egg','barrel','monument','statue','chest','pyre']): return 'Object'
    if 'boss' in (' '.join(subs)).lower(): return 'Boss'
    return 'Enemy'

def main():
    args=[a for a in sys.argv[1:] if not a.startswith('--')]
    outdir = sys.argv[sys.argv.index('--out')+1] if '--out' in sys.argv else '.'
    data_dir = args[0] if args else find_data_dir('.')
    data, objs, le = parse_assets(os.path.join(data_dir,'sharedassets0.assets'))
    loc = Localizer(os.path.join(data_dir,'resources.assets'))
    rows=[]; seen=set()
    for ch in iter_characters(data, objs, le):
        if ch['internal'].startswith('DialogueCharacter'): continue
        if ch['internal'] in EXCLUDED_INTERNALS: continue
        if ch['internal'] in seen: continue
        seen.add(ch['internal'])
        subs=ch['subtypes']
        name=loc.resolve(ch['nameKey']) if ch['nameKey'] else ch['internal']
        kind=classify(ch['internal'], subs, ch['hp'], ch['atk'])
        sl=(' '.join(subs)).lower()
        rows.append(dict(name=name or ch['internal'], internal=ch['internal'], set=ch['set'],
            hp=ch['hp'] or 0, atk=ch['atk'] or 0, subtypes='; '.join(dict.fromkeys(subs)),
            kind=kind, is_boss=('boss' in sl), is_support=('support' in sl),
            is_snack=('snack' in sl)))
    rows.sort(key=lambda r:r['name'].lower())
    os.makedirs(outdir, exist_ok=True)
    with open(os.path.join(outdir,'roster.json'),'w') as f: json.dump(rows,f,indent=1)
    print(f"wrote roster.json — {len(rows)} characters (cutscene portraits excluded)")
    if '--no-xlsx' not in sys.argv:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb=Workbook(); ws=wb.active; ws.title='Roster'
            cols=[('Name','name',34),('Internal','internal',34),('Set','set',8),('Kind','kind',13),
                  ('HP','hp',7),('ATK','atk',7),('Subtypes','subtypes',30)]
            ws.append([c[0] for c in cols])
            for c in ws[1]: c.font=Font(name='Arial',bold=True,color='FFFFFF'); c.fill=PatternFill('solid',start_color='1F3864')
            for r in rows:
                ws.append([r[k] for _,k,_ in cols])
                for c in ws[ws.max_row]: c.font=Font(name='Arial')
            ws.freeze_panes='A2'; ws.auto_filter.ref=f"A1:G{ws.max_row}"
            for i,(_,_,w) in enumerate(cols,1): ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=w
            wb.save(os.path.join(outdir,'roster.xlsx')); print("wrote roster.xlsx")
        except ImportError:
            print("(openpyxl not installed; skipped roster.xlsx)")

if __name__=='__main__': main()
