#!/usr/bin/env python3
"""
mt2_build_outputs.py — combine roster.json + scaling.json into deliverables:
  * scaled_<difficulty>.xlsx  — enemy ATK & HP by visit order (O1-O4)
  * a filtered boss sheet (internal name prefix, default Boss_SoulSavior)
  * scaling_page.html          — interactive calculator + Overgrowth table

Soul Savior filter is ON by default (the campaign mode). Pass --all to include
every enemy in the game (roguelike, etc.).

Reads scaling.json so the FORMULAS update automatically when the game is repatched.
Run mt2_extract_roster.py / unitypy_extract_roster.py and the scaling script first
(same --out dir).

Usage:
  python3 mt2_build_outputs.py --in OUTDIR [--difficulty Overgrowth] [--prefix Boss_SoulSavior] [--all]
"""
import sys, os, re, json

def getarg(flag, default=None):
    return sys.argv[sys.argv.index(flag)+1] if flag in sys.argv else default

SOUL_SAVIOR_EXTRA = {
    'EnchanterT1_Speed',  # Energizing Flautist — base game enemy used in Soul Savior
}

def is_soul_savior(e):
    subs = e['subtypes'].split('; ')
    return (
        any(s.startswith('SoulSavior_') for s in subs if s)
        or 'SoulSavior' in e['internal']
        or e['internal'] in SOUL_SAVIOR_EXTRA
    )

# Maps internal name → in-game variant display name for bosses whose display name
# is shared across all variants (Maera/Thaddeus/Tivi/Lylith/Lifemother).
# TrainBosses and battle bosses already have unique display names; omit them.
VARIANT_NAMES = {
    # Maera the Dutiful
    'Boss_SoulSavior_R1_Dutiful_Armor':               'Eldest Scion',
    'Boss_SoulSavior_R1_Dutiful_AscendDescend':       'Sibling Hierarchy',
    'Boss_SoulSavior_R1_Dutiful_Burst':               'Stern Sister',
    # Thaddeus the Indulgent
    'Boss_SoulSavior_R2_Indulgent_GorgeOnSlay':       'Insatiable',
    'Boss_SoulSavior_R2_Indulgent_ReduceCapacityResolve': 'Train Chomper',
    'Boss_SoulSavior_R2_Indulgent_Titanskin':         'Thick Skinned',
    # Tivi the Unruly
    'Boss_SoulSavior_R3_Unruly_DuplicateEnemy':       'Duplicitous',
    'Boss_SoulSavior_R3_Unruly_Scourge':              'Prankster',
    'Boss_SoulSavior_R3_Unruly_SniperAdvance':        'Mischevious Child',
    # Lylith the Spurned
    'Boss_SoulSavior_R4_Spurned_DualismAt50':         'Plaguebringer',
    'Boss_SoulSavior_R4_Spurned_WitherbloomOnAction': 'Energy Vampire',
    'Boss_SoulSavior_R4_Spurned_ArmorPerDebuff':      'Inoculation',
    # Lifemother
    'Boss_SoulSavior_Final_Lifemother_Debuffer':      'The Corpseflower',
    'Boss_SoulSavior_Final_Lifemother_Infested':      'The Swarmhost',
    'Boss_SoulSavior_Final_Lifemother_HealReanimate': 'The Undying Bloom',
}

# ---- build the multiplier model from scaling.json ----
def build_model(scaling):
    """Return funcs giving cumulative % by (difficulty, region, role, stat)."""
    def diff_obj(tier): return scaling.get('SoulSavior_DifficultyTier'+tier, {})
    def rd(n): return scaling.get('SoulSavior_RunDistance'+str(n), {})
    def role_key(role): return 'Boss' if role=='Boss' else role
    def diff_pct(diff, role, stat):
        if diff=='Bloom': return 0
        o=diff_obj('2' if diff=='Tangle' else '3')
        # Difficulty mutators: SetA = ATK percentages, SetB = HP percentages (opposite to region mutators)
        s = o.get('setB_attack' if stat=='hp' else 'setA_health', {})
        return s.get(role_key(role), 0) or 0
    def region_pct(region, role, stat):
        tot=0
        for n in range(1, region):   # region 2 -> RD1, region 3 -> RD1+RD2, ...
            # Region mutators: SetA = HP percentages, SetB = ATK percentages
            s=rd(n).get('setA_health' if stat=='hp' else 'setB_attack', {})
            tot += s.get(role_key(role), 0) or 0
        return tot
    def total_pct(diff, region, role, stat):
        return diff_pct(diff,role,stat) + region_pct(region,role,stat)
    return total_pct

import math

def rnd(x):
    return math.floor(x + 0.5)  # round-half-up (not banker's, not ceil)

def scaled(base, pct):
    return rnd(base * (1 + pct / 100))

def boss_overgrowth_scaled(base_atk, base_hp):
    """Verified incremental formula for Overgrowth O1-O4 bosses."""
    o1a = rnd(base_atk * 1.5)
    o1h = rnd(base_hp  * 1.65)
    o2a = o1a + 13
    o2h = o1h + rnd(base_hp * 0.10) + 1224
    o3a = o2a + 15
    o3h = o2h + rnd(base_hp * 0.10) + 2703
    o4a = o3a + 23
    o4h = o3h + rnd(base_hp * 0.145) + 3883
    return {'o1a': o1a, 'o1h': o1h, 'o2a': o2a, 'o2h': o2h,
            'o3a': o3a, 'o3h': o3h, 'o4a': o4a, 'o4h': o4h,
            'o5a': None, 'o5h': None}

def lifemother_overgrowth_scaled(base_atk, base_hp):
    """Lifemother final boss: O5 only, O1-O4 are blank."""
    return {'o1a': None, 'o1h': None, 'o2a': None, 'o2h': None,
            'o3a': None, 'o3h': None, 'o4a': None, 'o4h': None,
            'o5a': rnd(base_atk * 2.1) + 39,
            'o5h': rnd(base_hp  * 2.003) + 7794}

def enemy_overgrowth_scaled(base_atk, base_hp, is_support):
    """Empirical best-fit formula for non-boss Overgrowth enemies O1-O4.
    Game data (scaling.json) shows SetA=10%/SetB=4% per RunDistance, but observed
    values best fit 20% ATK / 7% HP additive from base. See docs/enemy_scaling.md.
    77% ATK match / 57% HP match against 35-enemy observation set.
    Support enemies receive no ATK scaling (Support absent from SetA in all mutators)."""
    out = {}
    for n in range(4):   # n=0 → O1, n=1 → O2, n=2 → O3, n=3 → O4
        o = n + 1
        out['o%da' % o] = base_atk if is_support else rnd(base_atk * (1.50 + n * 0.20))
        out['o%dh' % o] = math.ceil(base_hp * (1.35 + n * 0.07))
    out['o5a'] = out['o5h'] = None
    return out

# Bosses that only ever appear at O1 — O2-O5 are always blank.
O1_ONLY_BOSSES = {
    'Boss_SoulSavior_R0_TrainBoss_Defensive',  # Astrael the First Reborn
}

# Exact internal names to exclude from the boss sheet (duplicates/unused variants).
BOSS_EXCLUDED = {
    'Boss_SoulSavior_R1_TrainBoss_Athane_V2',
}

def is_lifemother(e):
    return 'Final_Lifemother' in e.get('internal', '')

def is_o1_only(e):
    return e.get('internal') in O1_ONLY_BOSSES

def role_of(e):
    if e['is_boss']: return 'Boss'
    if e['is_support']: return 'Support'
    return 'Heavy'   # health identical for all non-support combat roles; attack uses boss flag

def compute_orders(e, diff, total_pct):
    """Per-order ATK/HP dict (o1a..o5h, None where N/A) for one character.
    Single source of truth for both the xlsx/HTML build and the gamedata emitter."""
    role = role_of(e)
    if is_o1_only(e):
        if e['is_boss'] and diff == 'Overgrowth':
            r = boss_overgrowth_scaled(e['atk'], e['hp'])
            o1a, o1h = r['o1a'], r['o1h']
        else:
            o1a = scaled(e['atk'], total_pct(diff, 1, role, 'atk'))
            o1h = scaled(e['hp'],  total_pct(diff, 1, role, 'hp'))
        return {'o1a': o1a, 'o1h': o1h,
                'o2a': None, 'o2h': None, 'o3a': None, 'o3h': None,
                'o4a': None, 'o4h': None, 'o5a': None, 'o5h': None}
    if e['is_boss'] and diff == 'Overgrowth':
        if is_lifemother(e):
            return lifemother_overgrowth_scaled(e['atk'], e['hp'])
        return boss_overgrowth_scaled(e['atk'], e['hp'])
    if not e['is_boss'] and diff == 'Overgrowth':
        return enemy_overgrowth_scaled(e['atk'], e['hp'], e['is_support'])
    out = {}
    for o in range(1, 5):
        out['o%da' % o] = scaled(e['atk'], total_pct(diff, o, role, 'atk'))
        out['o%dh' % o] = scaled(e['hp'],  total_pct(diff, o, role, 'hp'))
    out['o5a'] = out['o5h'] = None
    return out

def main():
    indir=getarg('--in','.')
    diff=getarg('--difficulty','Overgrowth')
    prefix=getarg('--prefix','Boss_SoulSavior')
    soul_savior_only = '--all' not in sys.argv

    # Accept roster from either extractor (UnityPy or original)
    for fname in ('roster_unitypy.json', 'roster.json'):
        rpath = os.path.join(indir, fname)
        if os.path.exists(rpath):
            roster_full = json.load(open(rpath))
            break
    else:
        sys.exit(f"No roster JSON found in {indir}. Run an extraction script first.")

    for fname in ('scaling_unitypy.json', 'scaling.json'):
        spath = os.path.join(indir, fname)
        if os.path.exists(spath):
            scaling = json.load(open(spath))
            break
    else:
        sys.exit(f"No scaling JSON found in {indir}. Run the scaling extraction script first.")

    roster = [e for e in roster_full if is_soul_savior(e)] if soul_savior_only else roster_full
    mode_label = 'Soul Savior' if soul_savior_only else 'All enemies'
    print(f"Mode: {mode_label} — {len(roster)} of {len(roster_full)} characters")
    total_pct=build_model(scaling)

    def regions_for(e):
        return compute_orders(e, diff, total_pct)

    # ---- spreadsheets ----
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.page import PageMargins
        def sheet(wb, title, rows, onepage=False, with_regions=True):
            ws=wb.create_sheet(title) if wb.sheetnames!=['Sheet'] else wb.active
            ws.title=title
            if with_regions:
                hdr=['Name','Variant','Base ATK','Base HP','O1 ATK','O1 HP','O2 ATK','O2 HP','O3 ATK','O3 HP','O4 ATK','O4 HP','O5 ATK','O5 HP']
            else:
                hdr=['Name','Subtype','Base ATK','Base HP','O1 ATK','O1 HP','O2 ATK','O2 HP','O3 ATK','O3 HP','O4 ATK','O4 HP','O5 ATK','O5 HP']
            ws.append([f'{title}']); ws['A1'].font=Font(name='Arial',bold=True,size=12); ws.append([]); ws.append(hdr)
            for c in ws[3]:
                c.font=Font(name='Arial',bold=True,color='FFFFFF'); c.fill=PatternFill('solid',start_color='1F3864')
                c.alignment=Alignment(horizontal='center')
            for e in rows:
                variant = VARIANT_NAMES.get(e['internal'], '')
                r=regions_for(e)
                if with_regions:
                    ws.append([e['name'], variant, e['atk'], e['hp'], r['o1a'],r['o1h'],r['o2a'],r['o2h'],r['o3a'],r['o3h'],r['o4a'],r['o4h'],r['o5a'],r['o5h']])
                else:
                    subtype=e['subtypes'].split(';')[0].strip() if e['subtypes'] else ''
                    ws.append([e['name'], subtype, e['atk'], e['hp'], r['o1a'],r['o1h'],r['o2a'],r['o2h'],r['o3a'],r['o3h'],r['o4a'],r['o4h'],r['o5a'],r['o5h']])
                for j,c in enumerate(ws[ws.max_row]):
                    c.font=Font(name='Arial',size=10); c.alignment=Alignment(horizontal='left' if j<2 else 'right')
            ws.freeze_panes='A4'; ws.column_dimensions['A'].width=34
            ws.column_dimensions['B'].width=20
            for col in 'CDEFGHIJKLMN':
                ws.column_dimensions[col].width=9
            if onepage:
                ws.page_setup.orientation='landscape'; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
                ws.sheet_properties.pageSetUpPr.fitToPage=True
            return ws
        wb=Workbook()
        bosses=[e for e in roster if e['internal'].startswith(prefix) and e['internal'] not in BOSS_EXCLUDED]
        sheet(wb, f'Bosses ({diff})', bosses, onepage=True, with_regions=True)
        enemies=[e for e in roster if not e['is_boss']]
        sheet(wb, f'Enemies ({diff})', enemies, with_regions=False)

        # ---- Validation sheet (observed vs calculated) ----
        obs_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'boss_observations.csv')
        if os.path.exists(obs_path):
            import csv
            by_internal = {e['internal']: e for e in roster_full}
            variant_to_internal = {v: k for k, v in VARIANT_NAMES.items()}
            # also index by display name for TrainBosses (no VARIANT_NAMES entry)
            by_display = {e['name']: e for e in roster_full}
            order_num = {'O1':1,'O2':2,'O3':3,'O4':4}
            with open(obs_path, newline='') as f:
                obs_rows = [r for r in csv.DictReader(f) if r.get('obs_atk') or r.get('obs_hp')]
            if obs_rows:
                ws=wb.create_sheet('Validation')
                hdr=['Date','Boss','Variant','Order',
                     'Obs ATK','Calc ATK','ATK Diff',
                     'Obs HP','Calc HP','HP Diff','Notes']
                ws.append(hdr)
                for c in ws[1]:
                    c.font=Font(name='Arial',bold=True,color='FFFFFF')
                    c.fill=PatternFill('solid',start_color='1F3864')
                    c.alignment=Alignment(horizontal='center')
                for row in obs_rows:
                    internal = variant_to_internal.get(row['variant'])
                    e = by_internal.get(internal) or by_display.get(row['boss'])
                    on = order_num.get(row['order'].strip().upper(), 1)
                    if e:
                        r = regions_for(e)
                        calc_atk = r.get(f'o{on}a')
                        calc_hp  = r.get(f'o{on}h')
                    else:
                        calc_atk = calc_hp = '?'
                    obs_atk = int(row['obs_atk']) if row.get('obs_atk') else ''
                    obs_hp  = int(row['obs_hp'])  if row.get('obs_hp')  else ''
                    atk_diff = (obs_atk - calc_atk) if isinstance(obs_atk, int) and isinstance(calc_atk, int) else ''
                    hp_diff  = (obs_hp  - calc_hp)  if isinstance(obs_hp,  int) and isinstance(calc_hp,  int) else ''
                    ws.append([row['date'], row['boss'], row['variant'], row['order'],
                               obs_atk, calc_atk, atk_diff,
                               obs_hp,  calc_hp,  hp_diff, row.get('notes','')])
                    for j,c in enumerate(ws[ws.max_row]):
                        c.font=Font(name='Arial',size=10)
                        c.alignment=Alignment(horizontal='left' if j in (0,1,2,3,10) else 'right')
                        if j==6 and isinstance(atk_diff,int) and atk_diff!=0:
                            c.font=Font(name='Arial',size=10,color='FF4444')
                        if j==9 and isinstance(hp_diff,int) and hp_diff!=0:
                            c.font=Font(name='Arial',size=10,color='FF4444')
                ws.freeze_panes='A2'
                ws.column_dimensions['A'].width=12
                ws.column_dimensions['B'].width=26
                ws.column_dimensions['C'].width=20
                for col in 'DEFGHIJK': ws.column_dimensions[col].width=10
                ws.column_dimensions['K'].width=30
                print(f"  validation sheet: {len(obs_rows)} observations")

        suffix = '_SoulSavior' if soul_savior_only else ''
        outx=os.path.join(indir, f'scaled_{diff}{suffix}.xlsx'); wb.save(outx)
        print(f"wrote {outx}  ({len(bosses)} bosses, {len(enemies)} enemies)")
    except ImportError:
        print("(openpyxl missing; skipped xlsx)")

    # ---- HTML page (calculator + table for chosen difficulty) ----
    enr=[]
    for e in roster:
        r=regions_for(e); enr.append({**{k:e[k] for k in ('name','set','hp','atk','is_boss','is_support','is_snack')}, **r})
    html=PAGE.replace('__DIFF__',diff).replace('__DATA__',json.dumps(enr,separators=(',',':'))).replace('__SCALING__',json.dumps(scaling,separators=(',',':')))
    outh=os.path.join(indir,'scaling_page.html'); open(outh,'w').write(html); print(f"wrote {outh}")

PAGE = r'''<!DOCTYPE html><html><head><meta charset="utf-8"><title>MT2 Scaling (__DIFF__)</title>
<style>body{font:14px sans-serif;background:#15171c;color:#e7e9ee;padding:20px}h1{font-size:20px}
table{border-collapse:collapse;width:100%;font-size:13px}th,td{padding:5px 8px;border-bottom:1px solid #2a2e38;text-align:right}
th{position:sticky;top:0;background:#1e2128;cursor:pointer}td.n,th.n{text-align:left}.a{color:#ff9d6e}.h{color:#7ee0a0}
input,select{background:#252934;color:#e7e9ee;border:1px solid #333845;border-radius:6px;padding:6px}
.wrap{max-height:70vh;overflow:auto;border:1px solid #333845;border-radius:8px;margin-top:10px}
.badge{font-size:10px;padding:1px 5px;border-radius:5px;margin-left:5px}.bb{background:#5a2b1d;color:#ffb38f}.bs{background:#1d3a5a;color:#8fc6ff}</style></head>
<body><h1>Monster Train 2 — Attack &amp; Health by visit order (__DIFF__)</h1>
<p>Generated from scaling.json. <input id="q" placeholder="filter name..."> <label><input type=checkbox id=hide> hide snacks/0-atk</label> <span id=c></span></p>
<div class="wrap"><table id=t><thead><tr><th class=n data-k=name>Enemy</th><th data-k=set>Set</th><th data-k=atk>Base ATK</th><th data-k=hp>Base HP</th>
<th class=a data-k=o1a>O1 ATK</th><th class=h data-k=o1h>O1 HP</th><th class=a data-k=o2a>O2 ATK</th><th class=h data-k=o2h>O2 HP</th>
<th class=a data-k=o3a>O3 ATK</th><th class=h data-k=o3h>O3 HP</th><th class=a data-k=o4a>O4 ATK</th><th class=h data-k=o4h>O4 HP</th>
<th class=a data-k=o5a>O5 ATK</th><th class=h data-k=o5h>O5 HP</th></tr></thead><tbody id=b></tbody></table></div>
<script>const D=__DATA__;let k='name',d=1;
function render(){let q=q1.value.toLowerCase(),h=hide.checked,r=D.filter(e=>e.name.toLowerCase().includes(q));if(h)r=r.filter(e=>!(e.is_snack||e.atk===0));
r.sort((a,b)=>{let x=a[k],y=b[k];if(typeof x=='string'){x=x.toLowerCase();y=y.toLowerCase()}return(x>y?1:x<y?-1:0)*d});
b.innerHTML=r.map(e=>`<tr><td class=n>${e.name}${e.is_boss?'<span class="badge bb">BOSS</span>':''}${e.is_support?'<span class="badge bs">SUP</span>':''}</td><td>${e.set||''}</td><td class=a>${e.atk}</td><td class=h>${e.hp}</td>`+[1,2,3,4,5].map(g=>`<td class=a>${e['o'+g+'a']??''}</td><td class=h>${e['o'+g+'h']??''}</td>`).join('')+`</tr>`).join('');c.textContent=r.length+' enemies';}
const q1=document.getElementById('q'),hide=document.getElementById('hide'),b=document.getElementById('b'),c=document.getElementById('c');
document.querySelectorAll('#t th').forEach(t=>t.onclick=()=>{let kk=t.dataset.k;if(kk==k)d*=-1;else{k=kk;d=1}render()});q1.oninput=render;hide.onchange=render;render();</script></body></html>'''

if __name__=='__main__': main()
