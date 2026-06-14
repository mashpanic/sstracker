#!/usr/bin/env python3
"""
unitypy_extract_roster.py — UnityPy-based replacement for mt2_extract_roster.py.

Uses UnityPy for SerializedFile parsing and object enumeration; falls back to
positional byte reads for MonoBehaviour fields (no type trees in this build).
Output is identical to mt2_extract_roster.py: roster.json + roster.xlsx.

Usage:  python3 unitypy_extract_roster.py [DATA_DIR] [--out OUTDIR] [--no-xlsx]
"""
import sys, os, re, json
import UnityPy

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mt2_lib import Localizer, find_data_dir, STAT_HP_INDEX, STAT_ATK_INDEX, EXCLUDED_INTERNALS

TARGET = b"CharacterData_nameKey"

# MonoBehaviour header layout (no type tree, standard Unity 2022 Mono):
#   m_GameObject PPtr: i32 fileID + i64 pathID = 12 bytes
#   m_Enabled: u8 + 3-byte align pad = 4 bytes
#   m_Script PPtr: i32 fileID + i64 pathID = 12 bytes
#   m_Name: i32 length + chars                   <- starts at byte 28
_NAME_LEN_OFF = 28


def _read_name(raw: bytes) -> str:
    nlen = int.from_bytes(raw[_NAME_LEN_OFF:_NAME_LEN_OFF + 4], "little", signed=True)
    if nlen <= 0 or nlen > 500:
        return ""
    return raw[_NAME_LEN_OFF + 4: _NAME_LEN_OFF + 4 + nlen].decode("utf8", "replace")


def _read_stats(raw: bytes):
    """Return (hp, atk, statblock) from the int32 block after the prefab's 32-hex GUID."""
    pi = raw.find(b".prefab")
    if pi < 0:
        return None, None, []
    g = re.search(rb"[0-9a-f]{32}", raw[pi:])
    if not g:
        return None, None, []
    base = pi + g.start() + 32
    base = (base + 3) & ~3
    statblock = [
        int.from_bytes(raw[base + 4 * k: base + 4 * k + 4], "little", signed=True)
        for k in range(16)
    ]
    hp  = statblock[STAT_HP_INDEX]  if len(statblock) > STAT_HP_INDEX  else None
    atk = statblock[STAT_ATK_INDEX] if len(statblock) > STAT_ATK_INDEX else None
    return hp, atk, statblock


def iter_characters(sa0_path: str):
    """Yield character dicts from sharedassets0.assets using UnityPy."""
    env = UnityPy.load(sa0_path)
    for obj in env.objects:
        if obj.type.name != "MonoBehaviour":
            continue
        raw = obj.get_raw_data()
        if not raw or TARGET not in raw or b".prefab" not in raw:
            continue

        iname = _read_name(raw)
        hp, atk, statblock = _read_stats(raw)

        strs = [m.group().decode("utf8", "replace") for m in re.finditer(rb"[\x20-\x7e]{4,}", raw)]
        subs = [s.replace("SubtypesData_", "") for s in strs if s.startswith("SubtypesData_") and s != "SubtypesData_None"]
        subs = [re.sub(r"_[0-9a-f]{6,}.*$", "", s) for s in subs]
        nameKey = next((s for s in strs if s.startswith("CharacterData_nameKey")), None)
        pm = re.search(r"Character_Prefabs/([A-Za-z0-9_]+)/", "\n".join(strs))

        yield dict(
            internal=iname, pid=obj.path_id,
            hp=hp, atk=atk, statblock=statblock,
            subtypes=subs, nameKey=nameKey,
            set=(pm.group(1) if pm else ""),
            strings=strs,
        )


def classify(internal, subs, hp, atk):
    s = (" ".join(subs) + " " + internal).lower()
    if internal.startswith("DialogueCharacter"):
        return "Cutscene"
    if "snack" in s:
        return "Snack/Pickup"
    if any(k in s for k in ["treasure", "gold", "egg", "barrel", "monument", "statue", "chest", "pyre"]):
        return "Object"
    if "boss" in (" ".join(subs)).lower():
        return "Boss"
    return "Enemy"


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    outdir = sys.argv[sys.argv.index("--out") + 1] if "--out" in sys.argv else "."
    data_dir = args[0] if args else find_data_dir(".")

    loc = Localizer(os.path.join(data_dir, "resources.assets"))

    rows = []
    seen = set()
    for ch in iter_characters(os.path.join(data_dir, "sharedassets0.assets")):
        if ch["internal"].startswith("DialogueCharacter"):
            continue
        if ch["internal"] in EXCLUDED_INTERNALS:
            continue
        if ch["internal"] in seen:
            continue
        seen.add(ch["internal"])

        subs = ch["subtypes"]
        name = loc.resolve(ch["nameKey"]) if ch["nameKey"] else ch["internal"]
        kind = classify(ch["internal"], subs, ch["hp"], ch["atk"])
        sl = (" ".join(subs)).lower()
        rows.append(dict(
            name=name or ch["internal"],
            internal=ch["internal"],
            set=ch["set"],
            hp=ch["hp"] or 0,
            atk=ch["atk"] or 0,
            subtypes="; ".join(dict.fromkeys(subs)),
            kind=kind,
            is_boss=("boss" in sl),
            is_support=("support" in sl),
            is_snack=("snack" in sl),
        ))

    rows.sort(key=lambda r: r["name"].lower())
    os.makedirs(outdir, exist_ok=True)

    with open(os.path.join(outdir, "roster_unitypy.json"), "w") as f:
        json.dump(rows, f, indent=1)
    print(f"wrote roster_unitypy.json — {len(rows)} characters (cutscene portraits excluded)")

    if "--no-xlsx" not in sys.argv:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "Roster"
            cols = [
                ("Name",     "name",     34),
                ("Internal", "internal", 34),
                ("Set",      "set",       8),
                ("Kind",     "kind",     13),
                ("HP",       "hp",        7),
                ("ATK",      "atk",       7),
                ("Subtypes", "subtypes", 30),
            ]
            ws.append([c[0] for c in cols])
            for c in ws[1]:
                c.font = Font(name="Arial", bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", start_color="1F3864")
            for r in rows:
                ws.append([r[k] for _, k, _ in cols])
                for c in ws[ws.max_row]:
                    c.font = Font(name="Arial")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:G{ws.max_row}"
            for i, (_, _, w) in enumerate(cols, 1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
            wb.save(os.path.join(outdir, "roster_unitypy.xlsx"))
            print("wrote roster_unitypy.xlsx")
        except ImportError:
            print("(openpyxl not installed; skipped roster_unitypy.xlsx)")


if __name__ == "__main__":
    main()
